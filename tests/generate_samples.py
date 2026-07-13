"""Generate sample export files (CSV, JSON, Excel) for the exports/ directory."""

from __future__ import annotations

import csv
import json
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

SAMPLE_DATA = [
    {
        "id": 1,
        "title": "Global Markets Rally as Inflation Data Cools",
        "source": "Reuters",
        "author": "Jane Smith",
        "category": "Business",
        "published_time": "2025-07-11T08:30:00Z",
        "url": "https://www.reuters.com/markets/rally-inflation-2025",
        "scraped_time": "2025-07-11T09:00:00Z",
    },
    {
        "id": 2,
        "title": "India Launches New Space Mission to Study the Sun",
        "source": "NDTV",
        "author": "Rahul Verma",
        "category": "Science",
        "published_time": "2025-07-11T07:15:00Z",
        "url": "https://www.ndtv.com/science/india-sun-mission-2025",
        "scraped_time": "2025-07-11T09:00:00Z",
    },
    {
        "id": 3,
        "title": "Premier League Transfer Window: Top 10 Deals So Far",
        "source": "BBC News",
        "author": "Tom Brown",
        "category": "Sports",
        "published_time": "2025-07-11T06:00:00Z",
        "url": "https://www.bbc.com/sport/football/transfers-2025",
        "scraped_time": "2025-07-11T09:01:00Z",
    },
    {
        "id": 4,
        "title": "Climate Summit 2025: Key Takeaways from Day One",
        "source": "Al Jazeera",
        "author": "",
        "category": "World",
        "published_time": "2025-07-10T22:00:00Z",
        "url": "https://www.aljazeera.com/climate-summit-day1",
        "scraped_time": "2025-07-11T09:02:00Z",
    },
    {
        "id": 5,
        "title": "AI Regulation Bill Passes Senate Committee",
        "source": "CNN",
        "author": "Sarah Lee",
        "category": "Technology",
        "published_time": "2025-07-11T05:45:00Z",
        "url": "https://edition.cnn.com/tech/ai-regulation-bill-2025",
        "scraped_time": "2025-07-11T09:02:30Z",
    },
    {
        "id": 6,
        "title": "Monsoon Update: Heavy Rain Expected in Mumbai This Week",
        "source": "The Hindu",
        "author": "Priya Nair",
        "category": "Weather",
        "published_time": "2025-07-11T04:30:00Z",
        "url": "https://www.thehindu.com/weather/mumbai-monsoon-july",
        "scraped_time": "2025-07-11T09:03:00Z",
    },
    {
        "id": 7,
        "title": "Stock Market Hits All-Time High on Tech Earnings",
        "source": "Indian Express",
        "author": "Amit Sharma",
        "category": "Business",
        "published_time": "2025-07-11T08:00:00Z",
        "url": "https://indianexpress.com/business/stock-market-high-july",
        "scraped_time": "2025-07-11T09:04:00Z",
    },
    {
        "id": 8,
        "title": "New COVID Variant Detected: What Experts Say",
        "source": "Times of India",
        "author": "",
        "category": "Health",
        "published_time": "2025-07-11T03:15:00Z",
        "url": "https://timesofindia.com/health/covid-variant-july-2025",
        "scraped_time": "2025-07-11T09:04:30Z",
    },
    {
        "id": 9,
        "title": "NASA Confirms Water Ice on Mars Surface",
        "source": "Reuters",
        "author": "Michael Chen",
        "category": "Science",
        "published_time": "2025-07-10T20:00:00Z",
        "url": "https://www.reuters.com/science/nasa-mars-ice-2025",
        "scraped_time": "2025-07-11T09:05:00Z",
    },
    {
        "id": 10,
        "title": "Bollywood Box Office: Weekend Collection Report",
        "source": "NDTV",
        "author": "Sneha Kapoor",
        "category": "Entertainment",
        "published_time": "2025-07-11T01:00:00Z",
        "url": "https://www.ndtv.com/entertainment/box-office-july-2025",
        "scraped_time": "2025-07-11T09:05:30Z",
    },
]

EXPORTS_DIR = PROJECT_ROOT / "exports"
EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def generate_csv() -> None:
    path = EXPORTS_DIR / "sample_headlines.csv"
    fields = list(SAMPLE_DATA[0].keys())
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(SAMPLE_DATA)
    print(f"  [OK] CSV  -> {path}  ({path.stat().st_size:,} bytes)")


def generate_json() -> None:
    path = EXPORTS_DIR / "sample_headlines.json"
    with open(path, "w", encoding="utf-8") as f:
        json.dump(SAMPLE_DATA, f, indent=2, ensure_ascii=False)
    print(f"  [OK] JSON -> {path}  ({path.stat().st_size:,} bytes)")


def generate_excel() -> None:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    path = EXPORTS_DIR / "sample_headlines.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Headlines"

    # Header style
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    headers = list(SAMPLE_DATA[0].keys())
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header.replace("_", " ").title())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, record in enumerate(SAMPLE_DATA, 2):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=record[key])
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    wb.save(path)
    print(f"  [OK] XLSX -> {path}  ({path.stat().st_size:,} bytes)")


def generate_sample_db() -> None:
    """Populate the main database with sample data."""
    from database import DatabaseManager
    from logger import setup_logging

    setup_logging()

    db = DatabaseManager()
    inserted, skipped = db.insert_many(SAMPLE_DATA)
    print(f"  [OK] DB   -> {db.path}  ({inserted} inserted, {skipped} skipped)")
    db.log_activity("sample_data_loaded", f"{inserted} sample headlines inserted")


if __name__ == "__main__":
    print("Generating sample export files...\n")
    generate_csv()
    generate_json()
    generate_excel()
    generate_sample_db()
    print("\nAll sample files generated successfully.")
