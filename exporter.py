from __future__ import annotations

import csv
import json
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional

from logger import get_logger
from settings import Settings

log = get_logger(__name__)

_EXPORT_COLUMNS = [
    "id", "title", "source", "author", "category",
    "published_time", "url", "scraped_time",
]


class DataExporter:
    def __init__(self, export_dir: Optional[Path] = None) -> None:
        cfg = Settings()
        self._export_dir: Path = export_dir or cfg.export_folder
        self._export_dir.mkdir(parents=True, exist_ok=True)

    def to_csv(self, headlines: List[Dict[str, Any]], filename: Optional[str] = None) -> Path:
        path = self._resolve_path(filename, ".csv")
        rows = self._prepare_rows(headlines)

        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            writer = csv.DictWriter(fh, fieldnames=_EXPORT_COLUMNS, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

        log.info("Exported %d rows to CSV: %s", len(rows), path)
        return path

    def to_json(self, headlines: List[Dict[str, Any]], filename: Optional[str] = None) -> Path:
        path = self._resolve_path(filename, ".json")
        rows = self._prepare_rows(headlines)

        with open(path, "w", encoding="utf-8") as fh:
            json.dump(rows, fh, indent=2, ensure_ascii=False)

        log.info("Exported %d rows to JSON: %s", len(rows), path)
        return path

    def to_excel(self, headlines: List[Dict[str, Any]], filename: Optional[str] = None) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

        path = self._resolve_path(filename, ".xlsx")
        rows = self._prepare_rows(headlines)

        wb = Workbook()
        ws = wb.active
        ws.title = "Headlines"

        hdr_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        hdr_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_idx, col_name in enumerate(_EXPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.replace("_", " ").title())
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin

        for row_idx, record in enumerate(rows, 2):
            for col_idx, col_name in enumerate(_EXPORT_COLUMNS, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, ""))
                cell.border = thin
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

        ws.freeze_panes = "A2"

        wb.save(path)
        log.info("Exported %d rows to Excel: %s", len(rows), path)
        return path

    def export_all(self, headlines: List[Dict[str, Any]], base_name: Optional[str] = None) -> Dict[str, Path]:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        base = base_name or f"headlines_{stamp}"
        return {
            "csv": self.to_csv(headlines, f"{base}.csv"),
            "json": self.to_json(headlines, f"{base}.json"),
            "excel": self.to_excel(headlines, f"{base}.xlsx"),
        }

    def _resolve_path(self, filename: Optional[str], extension: str) -> Path:
        if filename:
            name = filename if filename.endswith(extension) else filename + extension
        else:
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            name = f"headlines_{stamp}{extension}"
        return self._export_dir / name

    @staticmethod
    def _prepare_rows(headlines: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return [{col: row.get(col, "") for col in _EXPORT_COLUMNS} for row in headlines]
